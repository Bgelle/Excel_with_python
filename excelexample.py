# from openpyxl import load_workbook
# def open_workbook(path):
#    workbook = load_workbook(filename=path)
#    print(f'Worksheet names: {workbook.sheetnames}')
#    sheet = workbook.active
#    print(sheet)
#    print(f'The title of the Worksheet is: {sheet.title}')
# if __name__ == '__main__':
#    open_workbook('D:\\Book1.xlsx')
# import openpyxl module
import openpyxl

# Give the location of the file
path = "D:\\Book2.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row=2, column=2)

print(cell_obj.value)